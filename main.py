import tkinter as tk
from tkinter import ttk
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import matplotlib.image as mpimg
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import numpy as np
from meteostat import Point, Daily
from datetime import datetime
from PIL import Image, ImageTk 

# Ruta al archivo Excel
ruta_archivo = "data2.xlsx"  # Cambia por la ruta de tu archivo
hoja = "Hoja1"  # Cambia por el nombre de la hoja si es necesario

# Cargar el archivo y la hoja
wb = load_workbook(filename=ruta_archivo, data_only=True)
ws = wb[hoja]

# Leer las filas y convertirlas a una lista de listas (omitimos la primera fila)
datos = [[celda.value for celda in fila] for fila in ws.iter_rows(min_row=2)]

# Lista de nodos a buscar (excepto "nodo 8")
nodos_buscados = {f"nodo{i}" for i in range(1, 19)}  # Nodos del "nodo1" al "nodo18"
nodos_buscados.discard("Nodo 9")  # Eliminar "nodo8" del conjunto

# Lista para almacenar los resultados
nodos = []

# Recorrer los datos y agregar a nodos si el valor de la columna 8 coincide
for i in range(len(datos)):
    # Eliminar espacios, convertir a minúsculas y comparar
    nodo_actual = str(datos[i][7]).replace(" ", "").lower()  # Eliminar espacios y convertir a minúsculas
    if nodo_actual in nodos_buscados:
        nodos.append(datos[i])

# Diccionario para almacenar la suma de la columna 7 por cada nodo
suma_por_nodo = {}

# Diccionario para contar cuántas filas hemos procesado por cada nodo
contador_por_nodo = {}

# Recorrer toda la lista nodos
for fila in nodos:
    nodo_nombre = fila[7]  # Nombre del nodo (columna 8)
    
    # Inicializamos el contador y la suma si es la primera vez que encontramos este nodo
    if nodo_nombre not in contador_por_nodo:
        contador_por_nodo[nodo_nombre] = 0
        suma_por_nodo[nodo_nombre] = 0

    # Si estamos dentro de las primeras 4 filas para este nodo
    if contador_por_nodo[nodo_nombre] < 4:
        columna_7 = fila[6]  # Columna 7 (índice 6)
        if isinstance(columna_7, (int, float)):  # Verificar que la columna 7 tenga un valor numérico
            suma_por_nodo[nodo_nombre] += columna_7  # Sumar el valor a la suma de ese nodo
        
        # Incrementar el contador para este nodo
        contador_por_nodo[nodo_nombre] += 1

# Crear la lista con la información de cada nodo y la suma de las primeras 4 filas
resultado = [[nodo, suma] for nodo, suma in suma_por_nodo.items()]

# Función para obtener el color de fondo de la alerta según el valor del ICA
def obtener_color_fondo_alerta(ica):
    if 0 <= ica <= 50:
        return "green"  # Buena calidad del aire
    elif 51 <= ica <= 100:
        return "yellow"  # Moderada
    elif 101 <= ica <= 150:
        return "orange"  # Dañina a la salud para grupos sensibles
    elif 151 <= ica <= 200:
        return "red"  # Dañina a la salud
    elif 201 <= ica <= 300:
        return "purple"  # Muy dañina a la salud
    elif 301 <= ica <= 500:
        return "brown"  # Peligrosa
    else:
        return "white"  # En caso de que el ICA no esté dentro del rango

# Crear la ventana principal
root = tk.Tk()
root.title("App con Resultado de Nodos")
root.geometry("800x600")

# Crear un marco para contener el navbar
main_frame = ttk.Frame(root)
main_frame.pack(fill="both", expand=True)

# Crear el Notebook para manejar las ventanas (pestañas)
notebook = ttk.Notebook(main_frame)
notebook.pack(fill="both", expand=True)

# Crear las páginas para el Notebook
pagina1 = ttk.Frame(notebook)
notebook.add(pagina1, text="Página 1")

pagina2 = ttk.Frame(notebook)
notebook.add(pagina2, text="Página 2")

pagina3 = ttk.Frame(notebook)
notebook.add(pagina3, text="Página 3")

# Crear un Treeview para mostrar el resultado
treeview = ttk.Treeview(pagina1, columns=("Nodo", "Suma"), show="headings")
treeview.heading("Nodo", text="Nodo")
treeview.heading("Suma", text="Suma de las primeras 4 filas")

def obtener_datos_climaticos():
    # Definir la ubicación (por ejemplo, Ciudad de Cali)
    location = Point(3.4516, -76.5320)


    # Definir el rango de fechas (desde 2024-01-01 hasta hoy)
    start = datetime(2024, 11, 16)
    end = datetime.now()

    # Obtener los datos diarios de clima
    data = Daily(location, start, end)
    data = data.fetch()


    print(data)

    # Convertir el DataFrame a una lista de listas
    data_lista = data.reset_index().values.tolist()

 

    lista_limpia_clima = []

    for i in range(len(data_lista)):
        for j in range(len(data_lista[0])):      
           if j < 8 :
               lista_limpia_clima.append(data_lista[i][j])  
        

    return lista_limpia_clima

datos_climaticos = obtener_datos_climaticos();
print(datos_climaticos)


def mostrar_pagina1():
    # Limpiar los widgets previos en la pestaña
    iconos_tk = []  
    for widget in pagina1.winfo_children():
        widget.destroy()

    # Cargar la imagen de fondo
    ruta_imagen_fondo = "nubes.jpg"  # Cambia por la ruta de tu imagen de fondo
    try:
        imagen_fondo = Image.open(ruta_imagen_fondo)
        imagen_fondo = imagen_fondo.resize((800, 600))  # Ajustar al tamaño de la pestaña
        fondo_tk = ImageTk.PhotoImage(imagen_fondo)
    except FileNotFoundError:
        print(f"Error: No se encontró la imagen de fondo en '{ruta_imagen_fondo}'.")
        return

    # Crear un Canvas para mostrar el fondo
    canvas_fondo = tk.Canvas(pagina1, width=800, height=600)
    canvas_fondo.pack(fill="both", expand=True)


    # Colocar la imagen en el Canvas
    canvas_fondo.create_image(0, 0, anchor="nw", image=fondo_tk)


        
    ruta_icono2 = "info2.png"  # Cambia por la ruta de tu ícono
    ica_icon_image2 = Image.open(ruta_icono2)
    ica_icon_image2 = ica_icon_image2.resize((400,200))  # Ajustar el tamaño del ícono
    ica_icon_tk2 = ImageTk.PhotoImage(ica_icon_image2)
    ica_icon = canvas_fondo.create_image(600, 380, anchor="s", image=ica_icon_tk2)  # Ajusta la posición Y
    mostrar_pagina1.ica_icon_tk2 = ica_icon_tk2  # Evitar que sea recolectado por el recolector de basuraor de basura

    # Almacenar la referencia de la imagen para evitar que sea recolectada por el recolector de basura
    mostrar_pagina1.fondo_tk = fondo_tk

    # Crear un marco para la barra de colores del ICA
    ica_bar_frame = ttk.Frame(canvas_fondo)
    ica_bar_frame.place(relx=0.1, rely=0.7, relwidth=0.8)  # Ajustar posición y tamaño relativos

    # Etiqueta descriptiva para la barra de colores del ICA
    ica_bar_label = ttk.Label(ica_bar_frame, text="ESTADO DEL AIRE EN CALI" , font=("Arial", 16, "bold"))
    ica_bar_label.pack(side="top", pady=5)

    # Crear un Canvas para dibujar la barra de colores
    canvas_ica_bar = tk.Canvas(ica_bar_frame, width=500, height=80)  # Incrementar altura del Canvas
    canvas_ica_bar.pack()


# Mostrar los datos de la lista `datos_climaticos`
    datos2 = obtener_datos_climaticos()

    
    # Iniciar la posición X en 10 y la posición Y en 50 (puedes ajustar estas posiciones)
    x_position = 130
    y_position = 50
    y2_position = 150

    datos2.pop(0)
    datos2.pop(6)
    datos2.pop()
    datos2.pop()
    datos2.pop()
    datos2.pop()
    datos2.pop()
    datos2.pop()
    datos2.pop()
    datos2.pop()

 

   # Iterar sobre los datos y mostrarlos horizontalmente
    for i in range(len(datos2)):
        canvas_fondo.create_text(x_position, y_position, text=str(datos2[i]), font=("Arial", 16), fill="black", anchor="w")
        x_position += 100  # Incrementar la posición X para el siguiente dato (ajusta el valor para cambiar el espacio entre los textos)


    canvas_fondo.create_text(130, y2_position, text=str("T° Promedio"), font=("Arial", 16), fill="black", anchor="w")
    canvas_fondo.create_text(230, y2_position, text=str("T° Min"), font=("Arial", 16), fill="black", anchor="w")
    canvas_fondo.create_text(330, y2_position, text=str("T° Max"), font=("Arial", 16), fill="black", anchor="w")
    canvas_fondo.create_text(430, y2_position, text=str("Velocidad"), font=("Arial", 16), fill="black", anchor="w")
    canvas_fondo.create_text(530, y2_position, text=str("Direccion"), font=("Arial", 16), fill="black", anchor="w")
    canvas_fondo.create_text(630, y2_position, text=str(""), font=("Arial", 16), fill="black", anchor="w")



    # Dibujar la barra de colores (igual que en tu código)
    rangos_ica = [(0, 50, "green"),
                  (51, 100, "yellow"),
                  (101, 150, "orange"),
                  (151, 200, "red"),
                  (201, 300, "purple"),
                  (301, 500, "brown")]

    x_start = 0  # Posición inicial en el eje X
    for rango in rangos_ica:
        x_end = (rango[1] / 500) * 500  # Escalar al tamaño del canvas (500 px de ancho)
        canvas_ica_bar.create_rectangle(x_start, 10, x_end, 50, fill=rango[2], outline=rango[2])  # Altura duplicada
        x_start = x_end  # Actualizar la posición inicial para el siguiente segmento

    # Etiquetas para los límites del rango
    for rango in rangos_ica:
        label_x = (rango[0] / 500) * 500  # Posición X de la etiqueta
        canvas_ica_bar.create_text(label_x, 60, text=str(rango[0]), anchor="n", font=("Arial", 8))

    # Crear un ícono para representar el valor del ICA (similar a tu código)
    ruta_icono = "pulmones2.png"  # Cambia por la ruta de tu ícono
    try:
        ica_icon_image = Image.open(ruta_icono)
        ica_icon_image = ica_icon_image.resize((50, 50))  # Ajustar el tamaño del ícono
        ica_icon_tk = ImageTk.PhotoImage(ica_icon_image)
    except FileNotFoundError:
        print(f"Error: No se encontró el archivo de ícono en '{ruta_icono}'.")
        return

    # Colocar el ícono en el canvas (más abajo de la barra)
    ica_icon = canvas_ica_bar.create_image(0, 50, anchor="s", image=ica_icon_tk)  # Ajusta la posición Y
    mostrar_pagina1.ica_icon_tk = ica_icon_tk  # Evitar que sea recolectado por el recolector de basura

    # Función para actualizar la posición del ícono
    def actualizar_posicion_ica(valor_ica):
        # Escalar el valor del ICA a la posición en el canvas
        posicion_x = (valor_ica / 500) * 500
        canvas_ica_bar.coords(ica_icon, posicion_x, 50)  # Mantén la posición Y en 50

    # Actualizar la posición inicial del ícono
    actualizar_posicion_ica(120)  
    # Cambia este valor para probar con diferentes ICAs

    # Cargar y mostrar el ícono `icono1.png`
    ruta_icono1 = "icono1.png"  # Cambia por la ruta de tu ícono
    ruta_icono2 = "icono2.png"  # Cambia por la ruta de tu ícono
    ruta_icono3 = "icono3.png"  # Cambia por la ruta de tu ícono
    ruta_icono4 = "icono4.png"  # Cambia por la ruta de tu ícono
    ruta_icono5 = "icono5.png"  # Cambia por la ruta de tu ícono
    ruta_icono6 = "icono6.png"  # Cambia por la ruta de tu ícono
    try:
        icono_image1 = Image.open(ruta_icono1)
        icono_image2 = Image.open(ruta_icono2)
        icono_image3 = Image.open(ruta_icono3)
        icono_image4 = Image.open(ruta_icono4)
        icono_image5 = Image.open(ruta_icono5)
        icono_image6 = Image.open(ruta_icono6)

        icono_image1 = icono_image1.resize((50, 50))  # Ajustar el tamaño del ícono
        icono_tk1 = ImageTk.PhotoImage(icono_image1)
        
        
        icono_image2 = icono_image2.resize((50, 50))  # Ajustar el tamaño del ícono
        icono_tk2 = ImageTk.PhotoImage(icono_image2)
        
        
        icono_image3= icono_image3.resize((50, 50))  # Ajustar el tamaño del ícono
        icono_tk3 = ImageTk.PhotoImage(icono_image3)
        
        
        icono_image4 = icono_image4.resize((50, 50))  # Ajustar el tamaño del ícono
        icono_tk4 = ImageTk.PhotoImage(icono_image4)
        
        
        icono_image5 = icono_image5.resize((50, 50))  # Ajustar el tamaño del ícono
        icono_tk5 = ImageTk.PhotoImage(icono_image5)
        
        
        icono_image6 = icono_image6.resize((50, 50))  # Ajustar el tamaño del ícono
        icono_tk6 = ImageTk.PhotoImage(icono_image6)

        
        
    except FileNotFoundError:
        print(f"Error: No se encontró el archivo de ícono en '{ruta_icono1}'.")
        return

    # Colocar el ícono en el Canvas en una posición fija
    canvas_fondo.create_image(150, 100, anchor="center", image=icono_tk1)  # Ajusta las coordenadas (400, 300)
    mostrar_pagina1.icono_tk1 = icono_tk1  # Evitar que sea recolectado por el recolector de basura

    canvas_fondo.create_image(250, 100, anchor="center", image=icono_tk2)  # Ajusta las coordenadas (400, 300)
    mostrar_pagina1.icono_tk2 = icono_tk2  # Evitar que sea recolectado por el recolector de basura

    canvas_fondo.create_image(350, 100, anchor="center", image=icono_tk3)  # Ajusta las coordenadas (400, 300)
    mostrar_pagina1.icono_tk3 = icono_tk3  # Evitar que sea recolectado por el recolector de basura

    canvas_fondo.create_image(450, 100, anchor="center", image=icono_tk4)  # Ajusta las coordenadas (400, 300)
    mostrar_pagina1.icono_tk4 = icono_tk4  # Evitar que sea recolectado por el recolector de basura

    canvas_fondo.create_image(550, 100, anchor="center", image=icono_tk5)  # Ajusta las coordenadas (400, 300)
    mostrar_pagina1.icono_tk = icono_tk5  # Evitar que sea recolectado por el recolector de basura

    canvas_fondo.create_image(650, 100, anchor="center", image=icono_tk6)  # Ajusta las coordenadas (400, 300)
    mostrar_pagina1.icono_tk6 = icono_tk6  # Evitar que sea recolectado por el recolector de basura

def obtener_color_ica(ica):
    """Devuelve el color según el rango del ICA."""
    if 0 <= ica <= 50:
        return "green"  # Buena calidad
    elif 51 <= ica <= 100:
        return "yellow"  # Moderada
    elif 101 <= ica <= 150:
        return "orange"  # No saludable para grupos sensibles
    elif 151 <= ica <= 200:
        return "red"  # No saludable
    elif 201 <= ica <= 300:
        return "purple"  # Muy no saludable
    else:
        return "brown"  # Peligroso

def mostrar_pagina2(event=None):
    # Configurar la página 2
    global fondo_tk, fondo_tk2, fondo_tk3  # Guardar referencias globales para evitar que se borren
    titulo = tk.Label(
        pagina2,
        text="NIVELES DE CONTAMINACION POR LOCALIDAD",
        font=("Arial", 16, "bold"),
        bg="white"
    )
    titulo.place(x=300, y=200)  # Posición relativa al contenedor (pagina2)
    titulo.pack()

    # Crear el canvas para dibujar
    canvas = tk.Canvas(pagina2, width=800, height=600, bg="white")
    canvas.pack(fill="both", expand=True)

    # Fondo 2 (mapa)
    ruta_fondo2 = "mapa.png"
    imagen_fondo2 = Image.open(ruta_fondo2)
    imagen_fondo2 = imagen_fondo2.resize((800, 600))  # Ajustar al tamaño del canvas
    fondo_tk2 = ImageTk.PhotoImage(imagen_fondo2)
    canvas.create_image(0, 0, anchor="nw", image=fondo_tk2)

    # Coordenadas y valores ICA para los 18 nodos
    coordenadas = [
        (100, 100), (200, 150), (300, 200), (400, 250), (500, 300),
        (150, 400), (250, 450), (350, 500), (450, 550), (550, 600),
        (120, 350), (220, 400), (320, 450), (420, 500), (520, 550),
        (170, 250), (270, 300), (370, 350)
    ]
    valores_ica = [20, 80, 130, 180, 250, 320, 50, 75, 100, 150, 200, 90, 45, 300, 70, 110, 160, 210]

    # Dibujar los nodos
    radio = 20  # Radio de los círculos
    for (x, y), ica in zip(coordenadas, valores_ica):
        color = obtener_color_ica(ica)
        canvas.create_oval(x - radio, y - radio, x + radio, y + radio, fill=color, outline="black")
        canvas.create_text(x, y, text=str(ica), fill="white", font=("Arial", 10, "bold"))

    # Fondo 3 (información adicional)
    ruta_fondo3 = "tavlainfo.png"
    imagen_fondo3 = Image.open(ruta_fondo3)
    imagen_fondo3 = imagen_fondo3.resize((550, 500))  # Ajustar al tamaño del canvas
    fondo_tk3 = ImageTk.PhotoImage(imagen_fondo3)
    canvas.create_image(800, 0, anchor="nw", image=fondo_tk3)

    root.mainloop()

def crear_alerta(frame, texto, estilo, color_fondo):
    """
    Crea y coloca una alerta en el frame.
    """
    alerta = ttk.Label(frame, text=texto, style=estilo)
    alerta.grid(sticky="w", padx=10, pady=5)
    alerta.configure(background=color_fondo)

def mostrar_pagina3(event=None):
    # Ordenar los resultados de mayor a menor suma
    resultado.sort(key=lambda x: x[1], reverse=True)
    
    # Limpiar los widgets existentes en la página
    for widget in pagina3.winfo_children():
        widget.destroy()

    # Crear el frame para las alertas
    alertas_frame = ttk.Frame(pagina3)
    alertas_frame.pack(fill="both", expand=True)

    # Crear un Canvas para dibujar el fondo
    canvas = tk.Canvas(pagina3, width=800, height=600)
    canvas.pack(fill="both", expand=True)

    # Cargar la imagen de fondo
    ruta_fondo4 = "respirar.jpg"
    imagen_fondo4 = Image.open(ruta_fondo4)
    imagen_fondo4 = imagen_fondo4.resize((800, 600))  # Ajustar al tamaño del canvas
    fondo_tk4 = ImageTk.PhotoImage(imagen_fondo4)

    canvas.create_image(0, 0, anchor="nw", image=fondo_tk4)
    canvas.image = fondo_tk4  # Esto es importante para evitar que se pierda la imagen

    # Ahora, puedes agregar otros elementos (como etiquetas, gráficos, etc.) encima de la imagen
    canvas.create_text(400, 50, text="NIVELES DE CONTAMINACION POR LOCALIDAD", font=("Arial", 16, "bold"), fill="white")

    # Estilo para las alertas
    style = ttk.Style()
    style.configure("Alerta.TLabel", font=("Arial", 12))

    # Títulos de las secciones
    label_mayor_contaminacion = ttk.Label(alertas_frame, text="EVITA SALIR AL AIRE LIBRE EN ESTAS LOCALIDADES", font=("Arial", 14, "bold"))
    label_mayor_contaminacion.grid(row=0, column=0, sticky="w", padx=10, pady=10)

    # Mostrar los 5 nodos con mayor contaminación
    for i, (nodo, suma) in enumerate(resultado[:5]):
        print(suma)
        if nodo.rstrip() == "Nodo 1":
            nodo = "Eco Parque Corazon de Pance"
        elif nodo.rstrip() == "Nodo 2":
            nodo = "Eco Parque las Garzas"
        elif nodo.rstrip() == "Nodo 3":
            nodo = "Universidad San Buena"
        elif nodo.rstrip() == "Nodo 4":
            nodo = "Universidad ICESI"
        elif nodo.rstrip() == "Nodo 5":
            nodo = "Univerisdad Autonoma"
        elif nodo.rstrip() == "Nodo 6":
            nodo = "Universidad Javeriana"
        elif nodo.rstrip() == "Nodo 7":
            nodo = "Fundacion Universitaria SM"
        elif nodo.rstrip() == "Nodo 8":
            nodo = "Universidad Libre"
        elif nodo.rstrip() == "Nodo 9":
            nodo = "Colegio Bennett"
        elif nodo.rstrip() == "Nodo 10":
            nodo = "Colegio Nuevo Cambridge"
        elif nodo.rstrip() == "Nodo 11":
            nodo = "Club campestre Cali"
        elif nodo.rstrip() == "Nodo 12":
            nodo = "Universidad Catolica Melendez"
        elif nodo.rstrip() == "Nodo 13":
            nodo = "Holguines Trade Center"
        elif nodo.rstrip() == "Nodo 14":
            nodo = "Universidad Santiago Cali"
        elif nodo.rstrip() == "Nodo 15":
            nodo = "Zona America"
        elif nodo.rstrip() == "Nodo 16":
            nodo = "Fundacion Valle Lili"
        elif nodo.rstrip() == "Nodo 17":
            nodo = "Colegio nuestra señora del rosario"
        elif nodo.rstrip() == "Nodo 18":
            nodo = "Condominio Bagatelle"

        texto = f"{i + 1}. {nodo} - Contaminación: {suma}"

        # Establecer color según el rango
        if 0 <= suma <= 50:
            color_fondo = "green"
        elif 51 <= suma <= 100:
            color_fondo = "yellow"
        elif 101 <= suma <= 150:
            color_fondo = "orange"
        elif 151 <= suma <= 200:
            color_fondo = "red"
        elif 201 <= suma <= 300:
            color_fondo = "purple"
        elif 301 <= suma <= 500:
            color_fondo = "brown"

        # Crear alerta
        crear_alerta(alertas_frame, texto, "Alerta.TLabel", color_fondo)

    # Títulos de las secciones
    label_menor_contaminacion = ttk.Label(alertas_frame, text="ESTOS SITIOS SON PERFECTOS PARA SALIR", font=("Arial", 14, "bold"))
    label_menor_contaminacion.grid(row=6, column=0, sticky="w", padx=10, pady=10)

    # Mostrar los 5 nodos con menor contaminación
    for i, (nodo, suma) in enumerate(resultado[-5:]):

        if nodo.rstrip() == "Nodo 1":
            nodo = "Eco Parque Corazon de Pance"
        elif nodo.rstrip() == "Nodo 2":
            nodo = "Eco Parque las Garzas"
        elif nodo.rstrip() == "Nodo 3":
            nodo = "Universidad San Buena"
        elif nodo.rstrip() == "Nodo 4":
            nodo = "Universidad ICESI"
        elif nodo.rstrip() == "Nodo 5":
            nodo = "Univerisdad Autonoma"
        elif nodo.rstrip() == "Nodo 6":
            nodo = "Universidad Javeriana"
        elif nodo.rstrip() == "Nodo 7":
            nodo = "Fundacion Universitaria SM"
        elif nodo.rstrip() == "Nodo 8":
            nodo = "Universidad Libre"
        elif nodo.rstrip() == "Nodo 9":
            nodo = "Colegio Bennett"
        elif nodo.rstrip() == "Nodo 10":
            nodo = "Colegio Nuevo Cambridge"
        elif nodo.rstrip() == "Nodo 11":
            nodo = "Club campestre Cali"
        elif nodo.rstrip() == "Nodo 12":
            nodo = "Universidad Catolica Melendez"
        elif nodo.rstrip() == "Nodo 13":
            nodo = "Holguines Trade Center"
        elif nodo.rstrip() == "Nodo 14":
            nodo = "Universidad Santiago Cali"
        elif nodo.rstrip() == "Nodo 15":
            nodo = "Zona America"
        elif nodo.rstrip() == "Nodo 16":
            nodo = "Fundacion Valle Lili"
        elif nodo.rstrip() == "Nodo 17":
            nodo = "Colegio nuestra señora del rosario"
        elif nodo.rstrip() == "Nodo 18":
            nodo = "Condominio Bagatelle"

        texto = f"{i + 1}. {nodo} - Contaminación: {suma}"

        # Establecer color según el rango
        if 0 <= suma <= 50:
            color_fondo = "green"
        elif 51 <= suma <= 100:
            color_fondo = "yellow"
        elif 101 <= suma <= 150:
            color_fondo = "orange"
        elif 151 <= suma <= 200:
            color_fondo = "red"
        elif 201 <= suma <= 300:
            color_fondo = "purple"
        elif 301 <= suma <= 500:
            color_fondo = "brown"

        # Crear alerta
        crear_alerta(alertas_frame, texto, "Alerta.TLabel", color_fondo)

# Vincular el cambio de pestaña para mostrar el contenido correcto
notebook.bind("<<NotebookTabChanged>>", lambda event: mostrar_pagina2(event) if notebook.index("current") == 1 else mostrar_pagina3(event) if notebook.index("current") == 2 else mostrar_pagina1())

# Iniciar con la página 1
mostrar_pagina1()

root.mainloop()